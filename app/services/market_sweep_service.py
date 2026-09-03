# StartWithA
# Copyright (C) 2024-2026 Kiran Mathews
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""
Market Sweep Service

Shared utilities for parsing CSV/Excel company files and seeding MarketSweep data.
Used by the Flask-Admin upload view and auto-seeded on app startup.
"""

import csv
import io
import logging
import os

import openpyxl

from app import db
from app.models.market_sweep import MarketSweep, MarketSweepCompany
from app.models.user import User
from app.services.sweep_link import link_sweep_row_by_isin
from app.utils.isin import is_valid_isin, normalize_isin

logger = logging.getLogger(__name__)

DEFAULT_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                                'data', 'market-sweeps')


def _norm_header(h):
    """Normalise header names: strip whitespace, lowercase, underscores."""
    if h is None:
        return ''
    return str(h).strip().lower().replace(' ', '_')


def parse_companies_file(file_or_path, filename=None):
    """Parse a CSV or Excel file and return a sorted list of company row dicts.

    Args:
        file_or_path: Either a file-like object (from Flask upload) or a
                      filesystem path string.
        filename:     Original filename (used to detect format). Required when
                      *file_or_path* is a file-like object. For path strings
                      the filename is derived automatically.

    Returns:
        List[dict] – rows sorted alphabetically by company_name, with
        normalised header keys.

    Raises:
        ValueError: If the file format is unsupported.
    """
    if filename is None:
        if isinstance(file_or_path, str):
            filename = file_or_path
        else:
            raise ValueError('filename is required when passing a file-like object')

    filename_lower = filename.lower()
    rows = []

    if filename_lower.endswith('.csv'):
        if isinstance(file_or_path, str):
            with open(file_or_path, 'r', encoding='utf-8-sig') as f:
                content = f.read()
        else:
            content = file_or_path.read().decode('utf-8-sig')

        reader = csv.DictReader(io.StringIO(content))
        rows = [{_norm_header(k): v for k, v in r.items()} for r in reader]

    elif filename_lower.endswith(('.xlsx', '.xls')):
        if isinstance(file_or_path, str):
            wb = openpyxl.load_workbook(file_or_path)
        else:
            wb = openpyxl.load_workbook(file_or_path)

        ws = wb.active
        headers = [_norm_header(cell.value) for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
    else:
        raise ValueError('Unsupported file format. Use CSV or Excel (.xlsx).')

    # The isin column must exist. Cells may be blank -- most rows have no ISIN
    # and never will -- but a file without the column is an old export, and
    # importing it would silently drop identity for every row it touches.
    if rows and 'isin' not in rows[0]:
        raise ValueError(
            'File must contain an "isin" column. Cells may be left blank, '
            'but the column is required.'
        )

    # Sort alphabetically by company_name
    rows.sort(key=lambda r: (r.get('company_name') or '').strip().lower())
    return rows


def upsert_sweep_companies(sweep_id, rows):
    """Update sweep rows in place, keyed on the id column. Never deletes.

    A sweep row is global but every user's MarketSweepDecision hangs off it by
    foreign key with a delete-orphan cascade. Deleting and recreating rows -- as
    the admin upload used to -- destroys every user's decisions on that sweep.
    Keying on id makes the operation non-destructive and survives renames.

    Rows present in the database but absent from the file are LEFT ALONE and
    counted. Removing a company from a sweep is a deliberate admin action, not a
    side effect of uploading a shorter file.

    Raises ValueError on a foreign id or an invalid ISIN so the caller can roll
    the whole upload back; a partial import is worse than none.
    """
    existing = {
        row.id: row
        for row in MarketSweepCompany.query.filter_by(sweep_id=sweep_id).all()
    }

    # A file carrying no ids is fine for a brand-new sweep -- every row is
    # necessarily new. But against a sweep that already has rows, such a file
    # can't be matched to anything: every row falls through to the insert path
    # and the sweep silently doubles, shadowing the rows that own the
    # decisions. Reject the whole upload instead.
    #
    # The test is the id VALUES, not the column: clearing the id cells of an
    # export leaves the header in place and does exactly the same damage. A
    # file that legitimately appends rows still carries the ids of the rows it
    # already had, so appending is unaffected.
    if existing and rows and not any(str(row.get('id') or '').strip() for row in rows):
        raise ValueError(
            f'This sweep already has {len(existing)} companies, but no row in '
            f'the file carries an "id". Uploading it would add duplicate rows '
            f'instead of updating the existing ones. Export the sweep first '
            f'so each row carries its id.'
        )

    seen = set()
    updated = inserted = 0
    isin_changed = []
    inserted_with_isin = []

    for index, row in enumerate(rows):
        name = str(row.get('company_name') or '').strip()
        if not name:
            continue

        isin = normalize_isin(row.get('isin'))
        if isin is not None and not is_valid_isin(isin):
            raise ValueError(
                f'Row {index + 1} ({name}): "{isin}" is not a valid ISIN '
                f'(12 characters, correct check digit).'
            )

        ticker = str(row.get('ticker') or '').strip() or None
        sector = str(row.get('sector') or '').strip() or None
        market_cap = str(row.get('market_cap') or '').strip() or None
        exchange = str(row.get('exchange') or '').strip() or None

        raw_id = str(row.get('id') or '').strip()
        if raw_id:
            try:
                row_id = int(float(raw_id))   # xlsx numeric cells arrive as 7.0
            except ValueError:
                raise ValueError(f'Row {index + 1} ({name}): id "{raw_id}" is not a number.')

            company = existing.get(row_id)
            if company is None:
                raise ValueError(
                    f'Row {index + 1} ({name}): id {row_id} does not belong to '
                    f'this sweep. This looks like the wrong file.'
                )

            company.company_name = name
            company.ticker = ticker
            company.sector_label = sector
            company.market_cap = market_cap
            company.exchange = exchange
            # A blank cell means "unknown", never "clear what is stored".
            if isin is not None and isin != company.isin:
                company.isin = isin
                isin_changed.append(company.id)
            company.sort_order = index

            seen.add(row_id)
            updated += 1
        else:
            new_company = MarketSweepCompany(
                sweep_id=sweep_id, company_name=name, ticker=ticker,
                isin=isin, sector_label=sector, market_cap=market_cap,
                exchange=exchange, sort_order=index,
            )
            db.session.add(new_company)
            if isin is not None:
                inserted_with_isin.append(new_company)
            inserted += 1

    # New rows only know their id once flushed, and the link pass needs those
    # ids. Flush once for the whole batch rather than per row: a first import
    # of thousands of rows would otherwise be thousands of single INSERTs.
    if inserted_with_isin:
        db.session.flush()
        isin_changed.extend(company.id for company in inserted_with_isin)

    return {
        'updated': updated,
        'inserted': inserted,
        'absent': len(set(existing) - seen),
        'errors': [],
        'isin_changed': isin_changed,
    }


def seed_market_sweeps(data_dir=None):
    """Auto-seed MarketSweep records from CSV/Excel files on startup.

    Scans *data_dir* for .csv / .xlsx files.  Each filename is treated as
    the country name (e.g. ``Australia.csv`` → country "Australia").

    **First run only.** A country already in the database is skipped, so
    editing a file on disk never changes an existing sweep. That is deliberate:
    sweep rows are shared and carry every user's decisions, and re-importing
    from disk on each boot would give a process restart authority over them.
    Changing an existing sweep goes through export → edit → upload in the admin,
    which matches rows by id and deletes nothing.

    Called once during app startup — safe to call repeatedly.
    """
    data_dir = data_dir or DEFAULT_DATA_DIR

    if not os.path.isdir(data_dir):
        return

    files = sorted(
        f for f in os.listdir(data_dir)
        if f.lower().endswith(('.csv', '.xlsx', '.xls'))
    )
    if not files:
        return

    admin_user = User.query.filter_by(is_admin=True).first()
    if not admin_user:
        logger.warning("seed_market_sweeps: no admin user found, skipping")
        return

    created = 0
    for filename in files:
        country = os.path.splitext(filename)[0]

        if MarketSweep.query.filter_by(country=country).first():
            continue

        filepath = os.path.join(data_dir, filename)
        try:
            rows = parse_companies_file(filepath)

            sweep = MarketSweep(
                name=f"{country} Market Sweep",
                country=country,
                description=f"Auto-seeded from {filename}",
                uploaded_by=admin_user.id,
            )
            db.session.add(sweep)
            db.session.flush()

            # The sweep is brand new here, so every row takes the insert path --
            # there's nothing yet for an id-less file to collide with. This is
            # also what makes ISIN validation happen for seeding, same as the
            # admin upload: build the rows through the shared upsert instead of
            # constructing MarketSweepCompany directly.
            upsert_sweep_companies(sweep.id, rows)

            sweep.total_companies = MarketSweepCompany.query.filter_by(
                sweep_id=sweep.id).count()
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            logger.error("seed_market_sweeps: error parsing %s — %s", filename, e)
            continue

        created += 1
        logger.info("seed_market_sweeps: created %s — %d companies",
                     country, sweep.total_companies)

    if created:
        logger.info("seed_market_sweeps: seeded %d new sweep(s)", created)


#: The columns an uploadable file must carry, in the order the parser
#: normalises them to. `id` is what makes a re-upload non-destructive: rows
#: are matched by identity rather than by name, so names may be edited freely.
EXPORT_COLUMNS = ['id', 'company_name', 'ticker', 'isin',
                  'sector', 'market_cap', 'exchange']


def build_sweep_export(sweep_id):
    """Serialise a sweep's companies as an .xlsx payload.

    The export is the supported source of an uploadable file. A hand-made file
    carries no ids, and the upsert refuses it rather than silently inserting
    duplicates alongside the rows that own the decisions.
    """
    companies = MarketSweepCompany.query.filter_by(
        sweep_id=sweep_id,
    ).order_by(MarketSweepCompany.sort_order).all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'companies'
    sheet.append(EXPORT_COLUMNS)

    for company in companies:
        sheet.append([
            company.id,
            company.company_name,
            company.ticker,
            company.isin,
            company.sector_label,
            company.market_cap,
            company.exchange,
        ])

    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def link_imported_isins(sweep_company_ids):
    """Link every row whose ISIN an import just set, for every owning user.

    An ISIN typed into a shared row answers "which company is this" for
    everyone who holds that security, not only for the admin who uploaded the
    file. Rows already linked are left alone.

    Returns the number of links created. Does not commit.
    """
    if not sweep_company_ids:
        return 0

    created = 0
    for row in MarketSweepCompany.query.filter(
        MarketSweepCompany.id.in_(sweep_company_ids)
    ).all():
        created += len(link_sweep_row_by_isin(row))

    return created
